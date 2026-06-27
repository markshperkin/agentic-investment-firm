import io

from openpyxl import Workbook
from openpyxl.styles import Font


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(r)
    return ws


def report_to_xlsx(report: dict, summary: dict) -> bytes:
    """Second reporting channel: a self-contained workbook (summary, holdings,
    trades, decision log, process metrics) built only from store numbers."""
    wb = Workbook()
    wb.remove(wb.active)

    m = report["metrics"]
    _sheet(wb, "Summary",
           ["Field", "Value"],
           [["Run", report["run_id"]], ["Date", report["replay_date"]],
            ["Headline", summary.get("headline", "")], ["Summary", summary.get("summary", "")],
            ["Starting cash", m["starting_cash"]], ["Cash", m["cash"]],
            ["Holdings value", m["holdings_value"]], ["Equity", m["equity"]],
            ["Portfolio return", m["portfolio_return"]],
            [f"{m['benchmark']} return", m["benchmark_return"]], ["Alpha", m["alpha"]],
            ["Filled trades", m["n_trades"]]])

    _sheet(wb, "Holdings",
           ["Ticker", "Qty", "Avg cost", "Mark", "Market value", "Unrealized P&L", "Realized P&L"],
           [[h["ticker"], h["quantity"], h["avg_cost_basis"], h["mark"], h["market_value"],
             h["unrealized_pnl"], h["realized_pnl"]] for h in report["holdings"]])

    _sheet(wb, "Trades",
           ["Ticker", "Side", "Qty", "Status", "Fill price", "Realized P&L", "As of"],
           [[t["ticker"], t["side"], t["quantity"], t["status"], t["fill_price"],
             t["realized_pnl"], t["as_of"]] for t in report["trades"]])

    _sheet(wb, "Decisions",
           ["Tick", "As of", "Ticker", "Path", "Action", "Shares", "Stance", "Confidence",
            "Thesis", "Citations"],
           [[d["tick_seq"], d["as_of"], d["ticker"], d["path"], d.get("action"),
             d.get("shares_held"), d["stance"], d["confidence"], d["thesis"],
             "; ".join(c.get("chunk_id") or "" for c in d["citations"])]
            for d in report["decisions"]])

    p = report["process"]
    _sheet(wb, "Process",
           ["Metric", "Value"],
           [["Total cost USD", p["total_cost_usd"]], ["Total tokens", p["total_tokens"]],
            ["Citation checks", p["citation_checks"]], ["Grounded views", p["grounded_views"]],
            ["Groundedness", p["groundedness"]], ["Refusals", p["refusals"]],
            ["Risk-engine rejects", p["risk_engine_rejects"]],
            ["Injection quarantines", p["injection_quarantines"]]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
